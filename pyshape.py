#coding:utf-8

"""
Donghui Chen, Wangmeng Song
May 15, 2017
"""

import math
import geopy.distance
import numpy as np
from scipy.spatial import cKDTree
from scipy import inf
from matplotlib import pyplot as plt
from scipy.cluster.hierarchy import dendrogram, linkage
from scipy.cluster.hierarchy import cophenet
from scipy.spatial.distance import pdist
from scipy.cluster.hierarchy import fcluster
from scipy.cluster.hierarchy import fclusterdata
import shapefile as sf
from descartes import PolygonPatch
from openpyxl import load_workbook
from shapely.geometry import Polygon, Point, MultiPolygon, LinearRing

# from datetime import datetime

# from mpl_toolkits.basemap import Basemap
# import numpy as np
# import matplotlib.pyplot as plt

# import sys
# sys.path.append('/System/Library/Frameworks/Python.framework/Versions/2.7/Extras/lib/python/mpl_toolkits')
# sys.path.append('/Library/Python/2.7/site-packages/mpl_toolkits')
# import axes_grid1


#定义一个点的经纬度（lat,lng）
class Location:
    def __init__(self, lat, lng):
        self.lat = lat
        self.lng = lng

def getListArrayDim(ainput, dim=0):
    """
    get the dimension of a list
    returns -1 if it is no list at all, 0 if list is empty
    and otherwise the dimensions of it
    """
    if isinstance(ainput, (list, np.ndarray)):
        if ainput == []:
            return dim
        dim = dim + 1
        dim = getListArrayDim(ainput[0], dim)
        return dim
    else:
        if dim == 0:
            return -1
        else:
            return dim

def BD2AMap(bdLat, bdLng):
    """
    Coordinate convertion: from BD to Gaode
    """
    x_pi = 3.14159265358979324 * 3000.0 / 180.0
    x = bdLng - 0.0065
    y = bdLat - 0.006
    z = math.sqrt(x * x + y * y) - 0.00002 * math.sin(y * x_pi)
    theta = math.atan2(y, x) - 0.000003 * math.cos(x * x_pi)
    AMapLng = z * math.cos(theta)
    AMapLat = z * math.sin(theta)
    return Location(AMapLat, AMapLng)

def calcDist(locationA, locationB):
    """
    input: two location coordinates: (latA, lngA) & (latB, lngB), could be tuple or list
    output: the distance of two location measuring in meters.
    """
    return geopy.distance.vincenty(locationA, locationB).km * 1000

def calcDistVec(destination, originVec):
    """
    destination: a list saving [lat, lng]
    originVec: a 2-dimensional array
    """
    if isinstance(originVec, (tuple, list)):
        if getListArrayDim(originVec) == 1:
            numLocation = 1
        else:
            numLocation = len(originVec)
    elif isinstance(originVec, np.ndarray):
        numLocation = originVec.shape[0]
    else:
        print "wrong input type in calcDistVec, originVec."
        return

    distVec = np.zeros([numLocation], dtype=int)
    for i in xrange(numLocation):
        distVec[i] = calcDist(destination, originVec[i])
    return distVec

# search neighborhood points with given center and radius, and sorted from small to large
def getNeighborhoodIdx(points, center, radius):
    # the distance of two points is defined as sqrt( (latA-latB)^2 + (lngA-lngB)^2 )
    # difference of 0.01 in lats and lngs is about 1.1km
    if len(points) == 1:
        neighborhoodIdx = [0]
        return neighborhoodIdx

    tree = cKDTree(points)
    neighborhoodIdx = [] # Put the neighbors of each point here
    distances, indices = tree.query(center, len(points), p=2, distance_upper_bound=radius)
    # print "getNeighborhoodIdx: ", indices, len(distances), distances

    for index, distance in zip(indices, distances):
        if distance == inf:
            break
        neighborhoodIdx.append(index)
    return neighborhoodIdx

# search neighborhood points with given center and radius, and sorted from small to large
# TBC
def getNeighborhoodIdx2(points, radius):
    # the distance of two points is defined as sqrt( (latA-latB)^2 + (lngA-lngB)^2 )
    # difference of 0.01 in lats and lngs is about 1.1km
    if len(points) == 1:
        neighborhoodIdx = [0]
        return neighborhoodIdx

    tree = cKDTree(points)
    neighborhoodIdx = [] # Put the neighbors of each point here
    print tree.query_ball_tree(tree, r=radius, p=2)
    # print "getNeighborhoodIdx: ", indices, len(distances), distances

def checkDistCondition(currentPoint, nextPoint, destination):
    # 检查下一个地点是否比现在的位置更靠近终点
    currentDist = calcDist(currentPoint, destination)
    nextDist = calcDist(nextPoint, destination)
    # print "checkDistCondition: ", currentDist, nextDist
    if currentDist >= nextDist:
        # print currentPoint, nextPoint
        return True
    else:
        return False

def angleBetweenVectorsDegrees(A, vertex, C):
    # https://stackoverflow.com/questions/42584259/python-code-to-calculate-angle-between-three-points-lat-long-coordinates
    """Return the angle between two vectors in any dimension space,
    in degrees."""
    # Convert the points to numpy latitude/longitude radians space
    a = np.radians(np.array(A))
    vertexR = np.radians(np.array(vertex))
    c = np.radians(np.array(C))
    # Vectors in latitude/longitude space
    sideA = a - vertexR
    sideC = c - vertexR
    # Adjust vectors for changed longitude scale at given latitude into 2D space
    lat = vertexR[0]
    sideA[1] *= math.cos(lat)
    sideC[1] *= math.cos(lat)
    return np.degrees(math.acos(np.dot(sideA, sideC) / (np.linalg.norm(sideA) * np.linalg.norm(sideC))))

def checkDirectionCondition(currentPoint, nextPoint, destination, maxAngle):
    # 检查下一个地点是否与现在的行驶方向顺路，顺路的定义是夹角小于maxAngle
    if angleBetweenVectorsDegrees(currentPoint, destination, nextPoint) < maxAngle:
        return True
    else:
        return False

def getSortedPointIdx(points, currentPoint, destination):
    # 获取当前点到所有订单地点的距离按照从小到大排序，并且下一个点需要更靠近终点
    MAXRADIUS = 1 # this corresponds to 110km, which is almost true for most of our case.
    MAXANGLE = 10
    neighborhoodIdx = getNeighborhoodIdx(points, currentPoint, MAXRADIUS)
    closestPointIdx = []
    for i in xrange(len(neighborhoodIdx)):
        if neighborhoodIdx[i] == inf:
            break
        if (checkDistCondition(currentPoint, points[neighborhoodIdx[i]], destination) and
            checkDirectionCondition(currentPoint, points[neighborhoodIdx[i]], destination, MAXANGLE)):
            closestPointIdx.append(neighborhoodIdx[i])

    return closestPointIdx

def getAllIndices(element, alist):
    """
    Find the index of an element in a list. The element can appear multiple times.
    input:  alist - a list
            element - objective element
    output: index of the element in the list
    """
    result = []
    offset = -1
    while True:
        try:
            offset = alist.index(element, offset + 1)
        except ValueError:
            return result
        result.append(offset)

def fancy_dendrogram(*args, **kwargs):
    max_d = kwargs.pop('max_d', None)
    if max_d and 'color_threshold' not in kwargs:
        kwargs['color_threshold'] = max_d
    annotate_above = kwargs.pop('annotate_above', 0)

    ddata = dendrogram(*args, **kwargs)

    if not kwargs.get('no_plot', False):
        plt.title('Hierarchical Clustering Dendrogram (truncated)')
        plt.xlabel('sample index or (cluster size)')
        plt.ylabel('distance')
        for i, d, c in zip(ddata['icoord'], ddata['dcoord'], ddata['color_list']):
            x = 0.5 * sum(i[1:3])
            y = d[1]
            if y > annotate_above:
                plt.plot(x, y, 'o', c=c)
                plt.annotate("%.3g" % y, (x, y), xytext=(0, -5),
                             textcoords='offset points',
                             va='top', ha='center')
        if max_d:
            plt.axhline(y=max_d, c='k')
    return ddata


def getShapefile(excelFileName, excelSheetName, plotFlag):

    # wb = load_workbook('MapLimitPoint.xlsx')
    wb = load_workbook(excelFileName)
    for i in range(len(excelSheetName)):
        ws = wb[excelSheetName[i]]
        print excelSheetName[i]
        pointsArea = []
        for row in ws.iter_rows(min_row=ws['A1'].value, min_col=ws['B1'].value,max_col=ws['C1'].value, max_row=ws['D1'].value):
            onePoint = []
            for cell in row:
                onePoint.append(cell.value)
            pointsArea.append(onePoint)

        # writing region boundary into a shape file
        print 'shapefiles/CDShapeFiles/'+excelSheetName[i]
        w = sf.Writer()
        w.poly(parts=[pointsArea])
        w.field('FIRST_FLD', 'C')
        w.save('shapefiles/CDShapeFiles/'+excelSheetName[i])

    # plot the shapefile
    if plotFlag == 1:
        fig = plt.figure()
        BLUE = '#6699cc'
        for i in range(len(excelSheetName)):
            polys = sf.Reader("shapefiles/CDShapeFiles/"+excelSheetName[i]+".shp")
            poly = polys.iterShapes().next().__geo_interface__
            ax = fig.gca()
            ax.add_patch(PolygonPatch(poly, fc=BLUE, ec=BLUE, alpha=0.5, zorder=2 ))
            ax.axis('scaled')
        plt.show()

    return


def main():

    # -------------------------
    # Get shapefile from excel file
    # excelSheetName = ['wholeArea','region1','region2','region3','region4','region5','region6']
    # excelSheetName = ['wholeArea','region1','region2','region3','region4','region5','region6', 'eastSide', 'westSide', 'westOutSide', 'eastOutSide', 'westPickUpArea', 'allPickUpArea', 'eastPickUpArea']
    # excelSheetName = ['westPickUpArea', 'allPickUpArea', 'eastPickUpArea']
    # excelFileName = 'MapLimitPoint.xlsx'
    excelFileName = 'MapLimitPoint0822.xlsx'
    excelSheetName = ['easttwoandfiveoutfile']
    getShapefile(excelFileName, excelSheetName, 1)


    # -------------------------
    # check if the input point is in a shapefile area

    # #[锦里，九眼桥，锦江宾馆, 仁恒置地, 桐梓林, 骡马寺]
    # currentPoint = [30.662447, 104.072469] # 天府广场
    # points = [[30.650817,104.056385], [30.645582,104.095192], [30.654087,104.072528],
    #             [30.658646,104.072563], [30.621274,104.073749], [30.672531,104.071962]]
    # destination = [30.599595,104.040745] # 交界点

    # point = Point(104.072469,30.662447) # 天府广场
    # point = Point(104.040745,30.599595) # keyPoint
    # point = Point(104.042779,30.620844)

    # start = datetime.now()
    # polys = sf.Reader("shapefiles/CDShapeFiles/wholeArea.shp")
    # polygon = polys.shapes()
    # shpfilePoints = []
    # for shape in polygon:
    #     shpfilePoints = shape.points
    # polygon = Polygon(shpfilePoints)
    # print datetime.now()-start
    # # point in polygon test
    # if polygon.contains(point):
    #     print 'inside'
    # else:
    #     print 'OUT'
    #     polExt = LinearRing(polygon.exterior.coords)
    #     d = polExt.project(point)
    #     p = polExt.interpolate(d)
    #     closest_point_coords = list(p.coords)[0]
    #     print calcDist((104.042779,30.620844), (closest_point_coords[0],closest_point_coords[1]))

    # print datetime.now()-start

    # plt.figure()
    # for shape in polys.shapeRecords():
    #     x = [i[0] for i in shape.points[:]]
    #     y = [i[1] for i in shape.points[:]]
    #     plt.plot(x,y)
    # plt.show()

    # -------------------------
    # check if the input point is in a shapefile area

    # #[锦里，九眼桥，锦江宾馆, 仁恒置地, 桐梓林, 骡马寺]
    # # currentPoint = [30.662447, 104.072469] # 天府广场
    # points = [[30.650817,104.056385], [30.645582,104.095192], [30.654087,104.072528],
    #             [30.658646,104.072563], [30.621274,104.073749], [30.672531,104.071962]]
    # destination = [30.599595,104.040745] # 交界点
    # search neighborhood points with given center and radius, and sorted from small to large
    # the distance of two points is defined as sqrt( (latA-latB)^2 + (lngA-lngB)^2 )
    # difference of 0.01 in lats and lngs is about 1.1km

    # tree = cKDTree(points)
    # neighborhoodIdx = [] # Put the neighbors of each point here
    # indices = tree.query_ball_point(points, 0.005)
    # print "getNeighborhoodIdx: "
    # for i in range(len(indices)):
    #     print indices[i]

if __name__ == '__main__':
    main()




