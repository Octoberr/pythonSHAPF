#coding:utf-8

"""
Donghui Chen, Wangmeng Song
May 15, 2017
"""

from openpyxl import load_workbook
from collections import OrderedDict
import json 


def main():

    wb = load_workbook('MapLimitPoint0822.xlsx', data_only=True)
    # sheetName = ['region1', 'region2', 'region3', 'region4', 'region5', 'region6']
    # sheetName = ['SecondRing', '1ring', 'wholeArea', 'metro1']
    # sheetName = [ 'easttwoandfiveoutfile']#, 'SecondRing']
    # sheetName = ['westPickUpArea', 'allPickUpArea', 'eastPickUpArea']
    # sheetName = ['eastOutSide','westOutSide']
    # sheetName = ['eastSide','westSide']
    ws = wb['easttwoandfiveoutfile']
    # for row in ws.iter_rows(min_row=ws['A1'].value, min_col=ws['B1'].value, max_col=ws['C1'].value,
    #                         max_row=ws['D1'].value):
    #     print row
    pointsArea = []
    for el in ws:
        tmpdic = {}
        tmpdic['idx'] = el[0].value
        tmpdic['lng'] = el[1].value
        tmpdic['lat'] = el[2].value
        pointsArea.append(tmpdic)
    # for i in range(len(sheetName)):
    #     ws = wb[sheetName[i]]
    #     for row in ws.iter_rows(min_row=ws['A1'].value, min_col=ws['B1'].value,max_col=ws['C1'].value, max_row=ws['D1'].value):
    #         onePoint = {'idx': [], 'lat': [], 'lng': []}
    #         onePointLst = []
    #         for cell in row:
    #             onePointLst.append(cell.value)
    #         onePoint['idx'].append(i)
    #         onePoint['lat'].append(onePointLst[1])
    #         onePoint['lng'].append(onePointLst[0])
    #         pointsArea.append(onePoint)
    # print pointsArea
     
    # Serialize the list of dicts to JSON
    j = json.dumps(pointsArea)

    # Write to file
    with open('data.json', 'w') as f:
        f.write(j)

        

if __name__ == '__main__':
    main()




